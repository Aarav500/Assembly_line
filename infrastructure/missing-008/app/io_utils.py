import os
import io
import json
from typing import Iterator, Optional, Tuple, List, Dict
import pandas as pd
import ijson
from openpyxl import Workbook, load_workbook


SUPPORTED_INPUT_FORMATS = {"csv", "json", "jsonl", "ndjson", "xlsx", "xls"}
SUPPORTED_OUTPUT_FORMATS = {"csv", "json", "jsonl", "xlsx"}


def ensure_dir(path: str):
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)


def detect_format(path: str) -> str:
    ext = os.path.splitext(path)[1].lower().strip(".")
    if ext in {"jsonl", "ndjson"}:
        return "jsonl"
    if ext in {"xlsx", "xls"}:
        return "xlsx"
    return ext


def count_rows(path: str, fmt: Optional[str] = None, json_array_pointer: str = "item") -> int:
    fmt = fmt or detect_format(path)
    if fmt == "csv":
        with open(path, "rb") as f:
            # count newline characters
            count = 0
            buf = f.read(1024 * 1024)
            while buf:
                count += buf.count(b"\n")
                buf = f.read(1024 * 1024)
        # subtract header if present
        return max(count - 1, 0)
    elif fmt == "jsonl":
        with open(path, "rb") as f:
            count = 0
            buf = f.read(1024 * 1024)
            while buf:
                count += buf.count(b"\n")
                buf = f.read(1024 * 1024)
        return count
    elif fmt == "json":
        total = 0
        with open(path, "rb") as f:
            for _ in ijson.items(f, json_array_pointer):
                total += 1
        return total
    elif fmt == "xlsx":
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        max_row = ws.max_row or 0
        wb.close()
        return max(max_row - 1, 0)
    else:
        raise ValueError(f"Unsupported format for counting: {fmt}")


def read_csv_chunks(path: str, chunksize: int, **kwargs) -> Iterator[pd.DataFrame]:
    kwargs = {"low_memory": False, **kwargs}
    for chunk in pd.read_csv(path, chunksize=chunksize, **kwargs):
        yield chunk


def read_jsonl_chunks(path: str, chunksize: int, **kwargs) -> Iterator[pd.DataFrame]:
    for chunk in pd.read_json(path, orient="records", lines=True, chunksize=chunksize, **kwargs):
        yield chunk


def read_json_array_chunks(path: str, chunksize: int, pointer: str = "item") -> Iterator[pd.DataFrame]:
    # Stream array elements into batches
    batch: List[Dict] = []
    with open(path, "rb") as f:
        for obj in ijson.items(f, pointer):
            batch.append(obj)
            if len(batch) >= chunksize:
                yield pd.DataFrame(batch)
                batch = []
    if batch:
        yield pd.DataFrame(batch)


def read_excel_chunks(path: str, chunksize: int) -> Iterator[pd.DataFrame]:
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        headers = next(rows)
    except StopIteration:
        wb.close()
        return
    headers = list(headers) if headers is not None else []
    batch: List[List] = []
    for row in rows:
        batch.append(list(row))
        if len(batch) >= chunksize:
            df = pd.DataFrame(batch, columns=headers)
            yield df
            batch = []
    if batch:
        yield pd.DataFrame(batch, columns=headers)
    wb.close()


def writer_csv(path: str, df: pd.DataFrame, header_written: bool):
    ensure_dir(path)
    mode = "a" if header_written else "w"
    df.to_csv(path, mode=mode, index=False, header=not header_written)


def writer_jsonl(path: str, df: pd.DataFrame, header_written: bool):
    ensure_dir(path)
    mode = "a" if header_written else "w"
    with open(path, mode, encoding="utf-8") as f:
        f.write(df.to_json(orient="records", lines=True, force_ascii=False))
        if not df.empty and not str(df.to_json(orient="records", lines=True)).endswith("\n"):
            f.write("\n")


def _excel_create_with_header(path: str, columns: List[str]):
    ensure_dir(path)
    wb = Workbook(write_only=False)
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(columns)
    wb.save(path)
    wb.close()


def writer_excel(path: str, df: pd.DataFrame, header_written: bool):
    if df.empty:
        return
    if not os.path.exists(path):
        _excel_create_with_header(path, list(df.columns))
        header_written = True
    wb = load_workbook(path)
    ws = wb.active
    for row in df.itertuples(index=False, name=None):
        ws.append(list(row))
    wb.save(path)
    wb.close()


def get_reader(path: str, fmt: Optional[str], chunksize: int, json_array_pointer: str = "item") -> Iterator[pd.DataFrame]:
    fmt = fmt or detect_format(path)
    if fmt == "csv":
        return read_csv_chunks(path, chunksize)
    if fmt == "jsonl":
        return read_jsonl_chunks(path, chunksize)
    if fmt == "json":
        return read_json_array_chunks(path, chunksize, pointer=json_array_pointer)
    if fmt == "xlsx":
        return read_excel_chunks(path, chunksize)
    raise ValueError(f"Unsupported input format: {fmt}")


def get_writer(fmt: str):
    if fmt == "csv":
        return writer_csv
    if fmt == "jsonl":
        return writer_jsonl
    if fmt == "xlsx":
        return writer_excel
    raise ValueError(f"Unsupported output format for chunked writer: {fmt}")


def normalize_output_path(output_path: str, fmt: str) -> str:
    base, ext = os.path.splitext(output_path)
    wanted_ext = {
        "csv": ".csv",
        "json": ".json",
        "jsonl": ".jsonl",
        "xlsx": ".xlsx",
    }[fmt]
    if ext.lower() != wanted_ext:
        return base + wanted_ext
    return output_path

